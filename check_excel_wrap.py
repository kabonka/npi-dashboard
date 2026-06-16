import openpyxl, re

wb = openpyxl.load_workbook('C:/npi-dashboard/NPI_vx/npi_dashboard.xlsx')
ws = wb.active

for row in range(2, min(30, ws.max_row+1)):
    s = ws.cell(row, 15).value
    h = ws.cell(row, 16).value
    if s and '\n' in s:
        print(f"Row {row} Status HAS newline: {repr(s[:150])}")
    if h and '\n' in h:
        print(f"Row {row} Highlight HAS newline: {repr(h[:150])}")

# Also show raw values (no newline) for reference
print("\n--- Sample raw Status values (no newline) ---")
count = 0
for row in range(2, ws.max_row+1):
    s = ws.cell(row, 15).value
    if s and '\n' not in s and re.search(r'\d+[)）]', s):
        print(f"Row {row}: {repr(s[:150])}")
        count += 1
        if count >= 5: break
