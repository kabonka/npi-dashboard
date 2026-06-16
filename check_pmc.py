from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime

wb = load_workbook('Spec總表.xlsx')
sheet = wb['PMC備料需求']
today = datetime.now().date()
yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')

target_stages = {'DVT', 'EVT', 'MVT'}

marked = 0
for row in range(2, sheet.max_row + 1):
    stage = sheet.cell(row=row, column=5).value
    if isinstance(stage, str):
        stage = stage.strip()
    if stage not in target_stages:
        continue

    f_val = sheet.cell(row=row, column=6).value
    g_val = sheet.cell(row=row, column=7).value

    f_empty = (f_val is None or (isinstance(f_val, str) and f_val.strip() == ''))
    g_empty = (g_val is None or (isinstance(g_val, str) and g_val.strip() == ''))

    if f_empty and g_empty:
        continue

    mark = False

    if not f_empty:
        f_date = None
        if isinstance(f_val, datetime):
            f_date = f_val
        elif isinstance(f_val, str):
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%m-%d-%Y']:
                try:
                    f_date = datetime.strptime(f_val.strip(), fmt)
                    break
                except ValueError:
                    pass
        if f_date and f_date.date() < today:
            i_val = sheet.cell(row=row, column=9).value
            if i_val is None or (isinstance(i_val, str) and i_val.strip() == ''):
                mark = True

    if not g_empty:
        g_date = None
        if isinstance(g_val, datetime):
            g_date = g_val
        elif isinstance(g_val, str):
            for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%m/%d/%Y', '%m-%d-%Y']:
                try:
                    g_date = datetime.strptime(g_val.strip(), fmt)
                    break
                except ValueError:
                    pass
        if g_date and g_date.date() < today:
            j_val = sheet.cell(row=row, column=10).value
            k_val = sheet.cell(row=row, column=11).value
            j_empty = (j_val is None or (isinstance(j_val, str) and j_val.strip() == ''))
            k_empty = (k_val is None or (isinstance(k_val, str) and k_val.strip() == ''))
            if j_empty or k_empty:
                mark = True

    if mark:
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=col).fill = yellow_fill
        marked += 1

wb.save('Spec總表.xlsx')
print(f'完成！共标记 {marked} 行（Stage=DVT/EVT/MVT，淡黄色底色）')
