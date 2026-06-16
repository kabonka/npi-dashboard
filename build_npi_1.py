"""
从 Excel 生成 NPI Dashboard 数据并嵌入 HTML。
用法: python build_npi.py
"""

import json
import os
import re
import datetime
import hashlib
import sys

# 关键依赖检查：openpyxl
print("[DEBUG] Checking imports...")
try:
    import openpyxl
except ImportError as e:
    print("ERROR: openpyxl not installed!")
    print("Please run: pip install openpyxl")
    print(f"Details: {e}")
    sys.stdout.flush()
    input("Press Enter to exit...")
    sys.exit(1)
print("[DEBUG] openpyxl OK")

try:
    import urllib.request
except ImportError:
    urllib = None

# ── 路径配置 ──
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(BASE_DIR, "Spec總表.xlsx")
HTML_PATH  = os.path.join(BASE_DIR, "npi_dashboard.html")
HTML_PATH2 = os.path.join(BASE_DIR, "npi_dashboard2.html")
JSON_PATH  = os.path.join(BASE_DIR, "npi_data.json")
SEARCH_HTML_PATH = os.path.join(BASE_DIR, "npi_search.html")

# GitHub Pages 上的 npi_dashboard.html URL（用于实时对比 MP）
GITHUB_HTML_URL = "https://kabonka.github.io/npi-dashboard/npi_dashboard.html"

# Excel Schedule sheet 列映射 (1-indexed, Row 1=表头)
COL = {
    "model": 1, "mkt": 2, "series": 3, "segment": 4,
    "cpu": 5, "gpu": 6, "npm": 7, "spm": 8,
    "stage": 9,
    "id_frozen": 10,
    "Kickoff": 11, "DVT-start": 12, "EVT-start": 13,
    "MVT-start": 14, "BTO ready": 15, "ATS-start": 16,
    "MP": 17,
    "status": 18, "highlight": 19,
}

# 日期列名（按顺序）
DATE_KEYS = ["Kickoff", "DVT-start", "EVT-start", "MVT-start", "BTO ready", "ATS-start", "MP"]


def fmt_date(val):
    """把 datetime / 字符串 / na 转成 YYYY/MM/DD 格式字符串，无效返回 None"""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y/%m/%d")
    if isinstance(val, str):
        val = val.strip()
        if not val or val.lower() == "na" or val.lower() == "n/a":
            return None
        # 尝试解析常见日期格式
        for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%e", "%m/%d/%Y"):
            try:
                dt = datetime.datetime.strptime(val, fmt)
                return dt.strftime("%Y/%m/%d")
            except ValueError:
                continue
        return val
    return None


def read_excel():
    """读取 Excel Schedule sheet 并返回 records 列表"""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Schedule"]

    records = []
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row_idx, COL["model"]).value
        if not model or not str(model).strip():
            continue

        stage = ws.cell(row_idx, COL["stage"]).value
        if not stage or not str(stage).strip():
            continue

        # 构造 dates 对象
        dates = {}
        mp_val = None
        for dk in DATE_KEYS:
            v = fmt_date(ws.cell(row_idx, COL[dk]).value)
            if v is not None:
                dates[dk] = v
                if dk == "MP":
                    mp_val = v

        # 如果没有日期则跳过
        if not dates:
            continue

        record = {
            "model": str(model).strip(),
            "mkt": str(ws.cell(row_idx, COL["mkt"]).value or "").strip(),
            "cpu": str(ws.cell(row_idx, COL["cpu"]).value or "").strip(),
            "gpu": str(ws.cell(row_idx, COL["gpu"]).value or "").strip(),
            "npm": str(ws.cell(row_idx, COL["npm"]).value or "").strip(),
            "stage": str(stage).strip(),
            "status": str(ws.cell(row_idx, COL["status"]).value or "").strip(),
            "highlight": str(ws.cell(row_idx, COL["highlight"]).value or "").strip(),
            "mp_sort": mp_val or "",
            "orig_idx": len(records),
            "dates": dates,
        }
        records.append(record)

    wb.close()
    return records


def fetch_github_records():
    """从 GitHub Pages 获取最新的 npi_dashboard.html，解析其中的 JSON 数据"""
    try:
        req = urllib.request.Request(
            GITHUB_HTML_URL,
            headers={"User-Agent": "Mozilla/5.0"}
        )
        with urllib.request.urlopen(req, timeout=8) as resp:
            html = resp.read().decode("utf-8")
        # 从 HTML 中解析 const DATA = { ... };
        match = re.search(r'const DATA = (\{[\s\S]*?\});', html)
        if not match:
            print("⚠️ 无法从 GitHub HTML 中解析 JSON 数据")
            return {}
        json_str = match.group(1)
        # 移除/替换非法控制字符（JSON 字符串中不允许的字符）
        json_str = json_str.replace('\t', ' ')  # Tab 替换为空格
        json_str = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', json_str)  # 移除其他控制字符
        data = json.loads(json_str)
        records = {(r.get("model",""), r.get("mkt","")): r for r in data.get("records", [])}
        print(f"已从 GitHub 获取 {len(records)} 条记录用于对比")
        return records
    except Exception as e:
        print(f"⚠️ 从 GitHub 获取数据失败: {e}")
        return {}


def detect_mp_changes(new_records):
    """对比 GitHub 上的最新 JSON，检测 MP 日期或 stage 有变动的记录，标记 mp_changed"""
    # 先从 GitHub 获取最新数据
    github_records = fetch_github_records()
    
    # 如果 GitHub 获取失败，回退到本地 JSON
    if not github_records:
        print("回退到本地 JSON 对比...")
        if not os.path.exists(JSON_PATH):
            return new_records
        try:
            with open(JSON_PATH, "r", encoding="utf-8") as f:
                old_data = json.load(f)
            github_records = {(r.get("model",""), r.get("mkt","")): r for r in old_data.get("records", [])}
        except Exception:
            return new_records

    changed_count = 0
    for r in new_records:
        key = (r.get("model", ""), r.get("mkt", ""))
        old = github_records.get(key)
        if not old:
            # 新增记录，不算 MP 变动
            continue
        old_mp = old.get("dates", {}).get("MP", "")
        new_mp = r.get("dates", {}).get("MP", "")
        old_stage = old.get("stage", "")
        new_stage = r.get("stage", "")
        if old_mp != new_mp or old_stage != new_stage:
            r["mp_changed"] = True
            changed_count += 1

    if changed_count > 0:
        print(f"与 GitHub 版本对比，检测到 {changed_count} 条 MP/Stage 变动")
    else:
        print("与 GitHub 版本对比，未检测到 MP/Stage 变动")

    return new_records


def inject_html(records):
    """把数据写入 HTML 的 const DATA = { ... } 部分"""
    with open(HTML_PATH, "r", encoding="utf-8") as f:
        html = f.read()

    # 把 records 序列化为 JSON
    data_obj = {"records": records, "buildTime": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")}
    json_str = json.dumps(data_obj, ensure_ascii=False, indent=2)

    # 替换 const DATA = { ... }; 块
    pattern = r"const DATA = \{[\s\S]*?\n\};"
    replacement = f"const DATA = {json_str};"

    new_html, count = re.subn(pattern, replacement, html, count=1)
    if count == 0:
        raise RuntimeError("未找到 'const DATA = { ... };' 块，请检查 HTML 文件")

    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(new_html)


def inject_html2(records):
    """把数据写入 npi_dashboard2.html"""
    with open(HTML_PATH2, "r", encoding="utf-8") as f:
        html = f.read()

    data_obj = {"records": records, "buildTime": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")}
    json_str = json.dumps(data_obj, ensure_ascii=False, indent=2)

    pattern = r"const DATA = \{[\s\S]*?\n\};"
    replacement = f"const DATA = {json_str};"

    new_html, count = re.subn(pattern, replacement, html, count=1)
    if count == 0:
        raise RuntimeError("未在 npi_dashboard2.html 中找到 'const DATA = { ... };' 块")

    with open(HTML_PATH2, "w", encoding="utf-8") as f:
        f.write(new_html)


def generate_xlsx(records):
    """生成带折叠分组的 Excel 文件（H~M 列默认折叠，子行默认折叠）"""
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NPI Dashboard"

    # 表头
    dk = ["Kickoff","DVT-start","EVT-start","MVT-start","BTO ready","ATS-start","MP"]
    headers = ["No.","Model","MKT Name","Stage","CPU","GPU","NPM"] + dk + ["Status","Highlight"]
    header_font = Font(bold=True, color="FFFFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(headers, 1):
        cell = ws.cell(1, i, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 列宽
    col_widths = [5,10,35,8,28,38,10,12,12,12,12,12,12,12,8,50]
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    # Stage 颜色
    stage_colors = {
        'Kickoff':'4472C4','Design':'00B0F0','DVT':'7030A0','EVT':'ED7D31',
        'MVT':'FF4444','ATS':'70AD47','Study':'FFC000','MP':'FF69B4',
        'BTO-':'00BCD4','ATS-':'66BB6A','Pending':'8899AA'
    }

    # 数据行（按 Model 分组）
    from collections import OrderedDict
    groups = OrderedDict()
    for rec in records:
        m = rec.get('model','')
        groups.setdefault(m, []).append(rec)

    row_idx = 2
    no = 1
    child_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

    for model, recs in groups.items():
        for i, rec in enumerate(recs):
            dates = rec.get('dates', {})
            vals = [no, rec.get('model',''), rec.get('mkt',''), rec.get('stage',''),
                    rec.get('cpu',''), rec.get('gpu',''), rec.get('npm','')]
            for k in dk:
                vals.append(dates.get(k, ''))
            vals.extend([rec.get('status',''), rec.get('highlight','')])

            for j, v in enumerate(vals):
                cell = ws.cell(row_idx, j+1, v)
                cell.alignment = Alignment(vertical="center")
                if j == 0:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if j == 3:  # Stage 列加颜色
                    color = stage_colors.get(str(v), '888888')
                    cell.font = Font(bold=True, color=color, size=10)

            is_parent = (i == 0)
            if is_parent:
                ws.cell(row_idx, 2).font = Font(bold=True, size=11)
                ws.row_dimensions[row_idx].height = 20
            else:
                ws.cell(row_idx, 2).font = Font(size=10, color="666666")
                ws.cell(row_idx, 2).alignment = Alignment(indent=2, vertical="center")
                # 子行折叠
                ws.row_dimensions[row_idx].outlineLevel = 1
                ws.row_dimensions[row_idx].hidden = True
                # 子行背景色
                for c in range(1, len(headers)+1):
                    ws.cell(row_idx, c).fill = child_fill

            no += 1
            row_idx += 1

    # H~M 列（8~13）折叠并默认隐藏
    for c in range(8, 14):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].outlineLevel = 1
        ws.column_dimensions[col_letter].hidden = True

    # 大纲属性
    ws.sheet_properties.outlinePr = openpyxl.worksheet.properties.Outline(summaryBelow=False)

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"

    # 保存
    xlsx_path = HTML_PATH.replace('.html', '.xlsx')
    wb.save(xlsx_path)
    print(f"Excel 已保存: {xlsx_path}")
    return xlsx_path


def build_search_html(records):
    """生成 NPI Search Dashboard 的自包含 HTML 文件"""
    data_obj = {"records": records, "buildTime": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")}
    json_str = json.dumps(data_obj, ensure_ascii=False, indent=2)

    html = f'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>NPI Search Dashboard</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
:root{{
  --bg:#0f1117;--bg2:#1a1d27;--bg3:#242836;
  --border:#2e3345;--text:#e0e0e0;--text2:#8a8fa8;
  --kickoff:#4472C4;--design:#00B0F0;--dvt:#7030A0;
  --evt:#ED7D31;--mvt:#FF4444;--ats:#70AD47;
  --study:#FFC000;--mp:#FF69B4;
}}
body{{background:var(--bg);color:var(--text);font-family:'Segoe UI',system-ui,sans-serif;min-height:100vh}}
/* ── 头部 ── */
.top-bar{{background:var(--bg2);border-bottom:1px solid var(--border);padding:14px 24px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
.top-bar h1{{font-size:20px;color:#fff;font-weight:600}}
.top-bar .stats{{font-size:14px;color:var(--text2)}}
/* ── 筛选区 ── */
.filter-title{{padding:12px 24px 0;font-size:16px;font-weight:600;color:#fff;letter-spacing:.5px;background:var(--bg2)}}
.filter-bar{{background:var(--bg2);border-bottom:1px solid var(--border);padding:16px 24px;display:flex;gap:24px}}
.filter-col{{flex:1;display:flex;flex-direction:column;gap:10px}}
.filter-col label{{font-size:14px;color:var(--text2);font-weight:600;margin-bottom:2px}}
.filter-row{{display:flex;align-items:center;gap:10px}}
.filter-bar input[type=text]{{padding:10px 14px;border:1px solid var(--border);border-radius:6px;background:var(--bg3);color:#fff;font-size:15px;outline:none;width:100%;flex:1}}
.filter-bar input[type=text]::placeholder{{color:var(--text2)}}
.filter-bar input[type=text]:focus{{border-color:#4472C4}}
/* 多选下拉 */
.ms-wrap{{position:relative;display:inline-block;width:100%}}
.ms-btn{{padding:10px 32px 10px 14px;border:1px solid var(--border);border-radius:6px;background:var(--bg3);color:#fff;font-size:15px;cursor:pointer;width:100%;text-align:left;user-select:none}}
.ms-btn::after{{content:'\\25BC';position:absolute;right:12px;top:50%;transform:translateY(-50%);font-size:10px;color:var(--text2)}}
.ms-dd{{display:none;position:absolute;top:100%;left:0;background:var(--bg3);border:1px solid var(--border);border-radius:6px;margin-top:4px;width:100%;min-width:220px;max-height:300px;overflow-y:auto;z-index:100;box-shadow:0 4px 12px rgba(0,0,0,.4)}}
.ms-dd.show{{display:block}}
.ms-dd .ms-all{{padding:8px 12px;font-size:14px;color:#4472C4;cursor:pointer;border-bottom:1px solid var(--border);position:sticky;top:0;background:var(--bg3)}}
.ms-dd label{{display:block;padding:7px 12px;font-size:14px;color:var(--text);cursor:pointer}}
.ms-dd label:hover{{background:rgba(68,114,196,.15)}}
.ms-dd input[type=checkbox]{{margin-right:8px;accent-color:#4472C4;width:16px;height:16px}}
/* ── 列表区 ── */
.list-wrap{{padding:0 24px 24px;overflow-x:auto}}
.list-table{{width:100%;border-collapse:collapse;font-size:14px;margin-top:0}}
.list-table thead th{{position:sticky;top:0;z-index:5;background:var(--bg2);color:var(--text2);font-size:12px;font-weight:600;text-transform:uppercase;padding:10px 12px;border-bottom:2px solid var(--border);text-align:left;white-space:nowrap}}
.list-table tbody tr{{cursor:pointer;transition:background .15s}}
.list-table tbody tr:hover{{background:rgba(68,114,196,.12)}}
.list-table tbody td{{padding:9px 12px;border-bottom:1px solid var(--border);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;max-width:300px}}
.list-table .col-model{{min-width:90px;max-width:110px}}
.list-table .col-mkt{{min-width:200px;max-width:340px}}
.list-table .col-stage{{min-width:80px;max-width:90px}}
.list-table .col-bto{{min-width:100px;max-width:110px}}
.list-table .col-mp{{min-width:100px;max-width:110px}}
.list-table .col-npm{{min-width:70px;max-width:100px}}
/* Stage badge */
.badge{{display:inline-block;padding:2px 8px;border-radius:3px;font-size:11px;font-weight:600;color:#fff}}
.badge-Kickoff{{background:var(--kickoff)}}.badge-Design{{background:var(--design)}}
.badge-DVT{{background:var(--dvt)}}.badge-EVT{{background:var(--evt)}}
.badge-MVT{{background:var(--mvt)}}.badge-ATS{{background:var(--ats)}}
.badge-Study{{background:var(--study);color:#333}}.badge-MP{{background:var(--mp)}}
.badge-m{{background:var(--mvt)}}.badge-M{{background:var(--mp)}}
/* MP 变动行高亮 */
.list-table tbody tr.mp-changed{{background:rgba(255,0,0,0.2)!important;border-left:3px solid #ff4444}}
.list-table tbody tr.mp-changed:hover{{background:rgba(255,0,0,0.3)!important}}
/* 空状态 */
.empty{{text-align:center;padding:60px 20px;color:var(--text2);font-size:14px}}
/* ── 统计栏 ── */
.stats-bar{{display:flex;gap:12px;padding:12px 24px;background:var(--bg2);border-bottom:1px solid var(--border);flex-wrap:wrap;align-items:center}}
.stat-item{{display:flex;align-items:center;gap:6px;background:var(--bg3);padding:6px 14px;border-radius:6px;border:1px solid var(--border)}}
.stat-label{{font-size:12px;color:var(--text2);font-weight:500}}
.stat-num{{font-size:16px;font-weight:700;min-width:20px;text-align:center}}
.stat-item.stat-model .stat-num{{color:#4472C4}}
.stat-item.stat-Design .stat-num{{color:#00B0F0}}
.stat-item.stat-DVT .stat-num{{color:#7030A0}}
.stat-item.stat-EVT .stat-num{{color:#ED7D31}}
.stat-item.stat-MVT .stat-num{{color:#FF4444}}
.stat-item.stat-ATS .stat-num{{color:#70AD47}}
.stat-item{{cursor:pointer;transition:background .15s,box-shadow .15s}}
.stat-item:hover{{background:rgba(68,114,196,.12)}}
.stat-item.active{{box-shadow:inset 0 0 0 2px currentColor}}
/* ── 详情模态框 ── */
.modal-overlay{{display:none;position:fixed;inset:0;background:rgba(0,0,0,.6);z-index:9000;justify-content:center;align-items:center;padding:16px}}
.modal-overlay.active{{display:flex}}
.modal-box{{background:var(--bg);border:1px solid var(--border);border-radius:10px;width:100%;max-width:560px;max-height:90vh;overflow-y:auto;position:relative;box-shadow:0 8px 32px rgba(0,0,0,.5)}}
.modal-close{{position:absolute;top:12px;right:16px;background:none;border:none;color:var(--text2);font-size:24px;cursor:pointer;line-height:1;z-index:10}}
.modal-close:hover{{color:#fff}}
.modal-body{{padding:24px}}
.modal-body .hdr{{margin-bottom:20px;padding-bottom:16px;border-bottom:1px solid var(--border)}}
.modal-body .hdr h2{{color:#fff;font-size:20px;font-weight:600;margin-bottom:4px}}
.modal-body .hdr .mkt{{color:var(--text2);font-size:13px}}
.modal-body .grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:20px}}
.modal-body .field{{background:var(--bg3);padding:12px;border-radius:6px;border:1px solid var(--border)}}
.modal-body .field-label{{font-size:10px;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px}}
.modal-body .field-value{{font-size:14px;color:var(--text);font-weight:500}}
.modal-body .section{{margin-bottom:16px}}
.modal-body .section-title{{font-size:11px;color:var(--text2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px}}
.modal-body .section-body{{background:var(--bg3);padding:12px;border-radius:6px;border:1px solid var(--border);font-size:13px;line-height:1.6;word-break:break-word}}
</style>
</head>
<body>
<!-- 主体 -->
<div id="mainApp">
  <div class="top-bar">
    <h1>NPI Statistic</h1>
    <span class="stats" id="statsInfo"></span>
  </div>
  <div class="stats-bar" id="statsBar">
    <div class="stat-item stat-model" onclick="onStatClick('model')"><span class="stat-label">Model</span><span class="stat-num" id="statModel">0</span></div>
    <div class="stat-item stat-Design" onclick="onStatClick('Design')"><span class="stat-label">Design</span><span class="stat-num" id="statDesign">0</span></div>
    <div class="stat-item stat-DVT" onclick="onStatClick('DVT')"><span class="stat-label">DVT</span><span class="stat-num" id="statDVT">0</span></div>
    <div class="stat-item stat-EVT" onclick="onStatClick('EVT')"><span class="stat-label">EVT</span><span class="stat-num" id="statEVT">0</span></div>
    <div class="stat-item stat-MVT" onclick="onStatClick('MVT')"><span class="stat-label">MVT</span><span class="stat-num" id="statMVT">0</span></div>
    <div class="stat-item stat-ATS" onclick="onStatClick('ATS')"><span class="stat-label">ATS</span><span class="stat-num" id="statATS">0</span></div>
  </div>
  <div class="filter-title">NPI Select and Search</div>
  <div class="filter-bar">
    <div class="filter-col">
      <label>Model</label>
      <div class="ms-wrap" id="modelMsWrap">
        <div class="ms-btn" id="modelMsBtn" onclick="toggleMs('model')">All Models</div>
        <div class="ms-dd" id="modelMsDd"></div>
      </div>
      <input type="text" id="searchModel" placeholder="Search Model..." oninput="onSearchChange()">
    </div>
    <div class="filter-col">
      <label>MKT</label>
      <div class="ms-wrap" id="mktMsWrap">
        <div class="ms-btn" id="mktMsBtn" onclick="toggleMs('mkt')">All MKT</div>
        <div class="ms-dd" id="mktMsDd"></div>
      </div>
      <input type="text" id="searchMkt" placeholder="Search MKT..." oninput="onSearchChange()">
    </div>
  </div>
  <div class="list-wrap">
    <table class="list-table">
      <thead><tr>
        <th class="col-model">Model</th>
        <th class="col-mkt">MKT Name</th>
        <th class="col-stage">Stage</th>
        <th class="col-bto">BTO Ready</th>
        <th class="col-mp">MP</th>
        <th class="col-npm">NPM</th>
      </tr></thead>
      <tbody id="listBody"></tbody>
    </table>
    <div class="empty" id="emptyMsg" style="display:none">No matching records found</div>
  </div>
</div>

<!-- 详情模态框 -->
<div class="modal-overlay" id="detailModal">
  <div class="modal-box">
    <button class="modal-close" onclick="closeDetail()">&times;</button>
    <div class="modal-body" id="detailBody"></div>
  </div>
</div>

<script>
// ── 数据 ──
const DATA = {json_str};

// ── 筛选状态 ──
let filterModelVals=new Set();
let filterMktVals=new Set();
let searchModelVal='';
let searchMktVal='';
let statStageFilter='';

// ── 多选下拉 ──
function getUniqueModels(){{return [...new Set(DATA.records.map(r=>r.model))].sort((a,b)=>a.localeCompare(b,undefined,{{numeric:true}}));}}
function getUniqueMkts(){{return [...new Set(DATA.records.map(r=>r.mkt.split(' ')[0]))].sort();}}

function buildMsDropdown(id){{
  const dd=document.getElementById(id+'MsDd');
  const items=id==='model'?getUniqueModels():getUniqueMkts();
  let h='<div class="ms-all" onclick="msToggleAll(\\''+id+'\\')">(Select All / None)</div>';
  items.forEach(v=>{{
    h+='<label><input type="checkbox" value="'+escAttr(v)+'" onchange="msChange(\\''+id+'\\')"> '+escHtml(v)+'</label>';
  }});
  dd.innerHTML=h;
}}

function toggleMs(id){{
  const dd=document.getElementById(id+'MsDd');
  document.querySelectorAll('.ms-dd.show').forEach(d=>{{if(d!==dd)d.classList.remove('show');}});
  dd.classList.toggle('show');
}}
document.addEventListener('click',function(e){{
  if(!e.target.closest('.ms-wrap'))document.querySelectorAll('.ms-dd.show').forEach(d=>d.classList.remove('show'));
}});

function msToggleAll(id){{
  const dd=document.getElementById(id+'MsDd');
  const cbs=dd.querySelectorAll('input[type=checkbox]');
  const allChecked=[...cbs].every(c=>c.checked);
  cbs.forEach(c=>c.checked=!allChecked);
  msChange(id);
}}

function msChange(id){{
  const dd=document.getElementById(id+'MsDd');
  const cbs=dd.querySelectorAll('input[type=checkbox]');
  const vals=new Set();
  cbs.forEach(c=>{{if(c.checked)vals.add(c.value);}});
  if(id==='model')filterModelVals=vals; else filterMktVals=vals;
  const btn=document.getElementById(id+'MsBtn');
  const total=id==='model'?getUniqueModels().length:getUniqueMkts().length;
  btn.textContent=vals.size===0?(id==='model'?'All Models':'All MKT'):vals.size+' selected';
  statStageFilter='';
  document.querySelectorAll('.stat-item').forEach(el=>el.classList.remove('active'));
  renderList();
}}

function escAttr(s){{return String(s).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}
function escHtml(s){{return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');}}

// ── 统计 ──
function updateStats(){{
  const activeStages=['Design','DVT','EVT','MVT','ATS','m','M'];
  const filtered=getFilteredRecords();
  // Model统计：active stage中Model去重
  const modelSet=new Set();
  let design=0,dvt=0,evt=0,mvt=0,ats=0;
  filtered.forEach(r=>{{
    const s=(r.stage||'').trim();
    if(activeStages.includes(s))modelSet.add(r.model);
    if(s==='Design')design++;
    if(s==='DVT')dvt++;
    if(s==='EVT')evt++;
    if(s==='MVT'||s==='m')mvt++;
    if(s==='ATS')ats++;
  }});
  document.getElementById('statModel').textContent=modelSet.size;
  document.getElementById('statDesign').textContent=design;
  document.getElementById('statDVT').textContent=dvt;
  document.getElementById('statEVT').textContent=evt;
  document.getElementById('statMVT').textContent=mvt;
  document.getElementById('statATS').textContent=ats;
}}

// ── 统计点击筛选 ──
function onStatClick(stage){{
  if(stage==='model'){{
    // Model 按钮：重置所有筛选，回到默认状态
    statStageFilter='';
    filterModelVals=new Set();
    filterMktVals=new Set();
    searchModelVal='';
    searchMktVal='';
    document.getElementById('searchModel').value='';
    document.getElementById('searchMkt').value='';
    document.getElementById('modelMsBtn').textContent='All Models';
    document.getElementById('mktMsBtn').textContent='All MKT';
    // 重置下拉复选框
    document.querySelectorAll('#modelMsDd input[type=checkbox],#mktMsDd input[type=checkbox]').forEach(c=>c.checked=false);
    document.querySelectorAll('.stat-item').forEach(el=>el.classList.remove('active'));
    renderList();
    return;
  }}
  if(statStageFilter===stage){{statStageFilter='';}}else{{statStageFilter=stage;}}
  // 更新 active 样式
  document.querySelectorAll('.stat-item').forEach(el=>el.classList.remove('active'));
  if(statStageFilter){{
    const cls='stat-'+statStageFilter;
    document.querySelector('.stat-item.'+cls).classList.add('active');
  }}
  renderList();
}}

// ── 搜索 ──
function onSearchChange(){{
  searchModelVal=document.getElementById('searchModel').value.trim();
  searchMktVal=document.getElementById('searchMkt').value.trim();
  statStageFilter='';
  document.querySelectorAll('.stat-item').forEach(el=>el.classList.remove('active'));
  renderList();
}}

// ── 过滤+排序 ──
function getFilteredRecords(){{
  const activeStages=['Design','DVT','EVT','MVT','ATS','m','M'];
  const listStages=['Design','DVT','EVT','MVT','ATS','m'];
  return DATA.records.filter(r=>{{
    // 列表只显示 Design/DVT/EVT/MVT/ATS（m=MVT别名）
    const stg=(r.stage||'').trim();
    if(!listStages.includes(stg))return false;
    if(filterModelVals.size>0&&!filterModelVals.has(r.model))return false;
    if(filterMktVals.size>0&&!filterMktVals.has(r.mkt.split(' ')[0]))return false;
    if(searchModelVal){{const q=searchModelVal.toLowerCase();if(!r.model.toLowerCase().includes(q))return false;}}
    if(searchMktVal){{const q=searchMktVal.toLowerCase();if(!r.mkt.toLowerCase().includes(q))return false;}}
    // 统计栏筛选
    if(statStageFilter==='MVT'){{
      if(stg!=='MVT'&&stg!=='m')return false;
    }} else if(statStageFilter){{
      if(stg!==statStageFilter)return false;
    }}
    return true;
  }});
}}
function getSortedRecords(recs){{
  const s=[...recs];
  s.sort((a,b)=>a.model.localeCompare(b.model,undefined,{{numeric:true}}));
  return s;
}}

// ── 渲染列表 ──
function renderList(){{
  const hasFilter=filterModelVals.size>0||filterMktVals.size>0||searchModelVal||searchMktVal||statStageFilter;
  const tbody=document.getElementById('listBody');
  const empty=document.getElementById('emptyMsg');
  const table=document.querySelector('.list-table');

  if(!hasFilter){{
    tbody.innerHTML='';
    empty.style.display='none';
    table.style.display='none';
    document.getElementById('statsInfo').textContent='Updated: '+DATA.buildTime;
    updateStats();
    return;
  }}

  const filtered=getFilteredRecords();
  const sorted=getSortedRecords(filtered);
  table.style.display='';
  document.getElementById('statsInfo').textContent='Updated: '+DATA.buildTime;
  updateStats();

  if(!sorted.length){{tbody.innerHTML='';empty.style.display='block';return;}}
  empty.style.display='none';

  let h='';
  sorted.forEach(r=>{{
    const stage=r.stage||'';
    const bto=r.dates?r.dates['BTO ready']||'':'';
    const mp=r.dates?r.dates['MP']||'':'';
    const cls='badge badge-'+stage.replace(/\\s/g,'');
    const rowClass=r.mp_changed?' class="mp-changed"':'';
    h+='<tr'+rowClass+' onclick="openDetail('+r.orig_idx+')">';
    h+='<td class="col-model">'+escHtml(r.model)+'</td>';
    h+='<td class="col-mkt" title="'+escAttr(r.mkt)+'">'+escHtml(r.mkt)+'</td>';
    h+='<td class="col-stage"><span class="'+cls+'">'+escHtml(stage)+'</span></td>';
    h+='<td class="col-bto">'+escHtml(bto)+'</td>';
    h+='<td class="col-mp">'+escHtml(mp)+'</td>';
    h+='<td class="col-npm">'+escHtml(r.npm||'')+'</td>';
    h+='</tr>';
  }});
  tbody.innerHTML=h;
}}

// ── 详情模态框 ──
function openDetail(idx){{
  const r=DATA.records.find(x=>x.orig_idx===idx);
  if(!r)return;
  document.getElementById('detailBody').innerHTML=generateDetailHTML(r);
  document.getElementById('detailModal').classList.add('active');
  document.body.style.overflow='hidden';
}}
function closeDetail(){{
  document.getElementById('detailModal').classList.remove('active');
  document.body.style.overflow='';
}}
document.getElementById('detailModal').addEventListener('click',function(e){{
  if(e.target===this)closeDetail();
}});

// ── 排版工具：遇「數字)」或「數字）」換行分段 ──
function formatSectionText(text){{
  if(!text)return 'N/A';
  if(!/\\d+[)）]/.test(text))return escHtml(text);
  // 在每個非開頭的「數字)」或「數字）」前斷行
  const formatted=text.replace(/([^\\n])(\\s*)(\\d+[)）])/g,'$1\\n$3');
  const parts=formatted.split('\\n').map(s=>s.trim()).filter(s=>s);
  if(parts.length<=1)return escHtml(text);
  return parts.map(p=>'<div style="margin-bottom:6px">'+escHtml(p)+'</div>').join('');
}}

function generateDetailHTML(r){{
  const stage=r.stage||'';
  const bto=r.dates?r.dates['BTO ready']||'':'';
  const mp=r.dates?r.dates['MP']||'':'';
  const cls='badge badge-'+stage.replace(/\\s/g,'');
  const dates=r.dates||{{}};
  const stageColors={{Kickoff:'#4472C4',Design:'#00B0F0',DVT:'#7030A0',EVT:'#ED7D31',MVT:'#FF4444',ATS:'#70AD47',Study:'#FFC000',MP:'#FF69B4',m:'#FF4444',M:'#FF69B4'}};
  const stageColor=stageColors[stage]||'#888';

  let timelineH='';
  const dk=['Kickoff','DVT-start','EVT-start','MVT-start','BTO ready','ATS-start','MP'];
  dk.forEach(k=>{{
    if(dates[k])timelineH+='<div style="display:flex;justify-content:space-between;padding:4px 0;border-bottom:1px solid #2e3345"><span style="color:#8a8fa8;font-size:12px">'+k+'</span><span style="color:#e0e0e0;font-size:12px">'+dates[k]+'</span></div>';
  }});

  let h='';
  h+='<div class="hdr"><h2>'+escHtml(r.model)+'</h2><div class="mkt">'+escHtml(r.mkt)+'</div></div>';
  h+='<div class="grid">';
  h+='<div class="field"><div class="field-label">Stage</div><div class="field-value"><span class="badge" style="background:'+stageColor+'">'+escHtml(stage)+'</span></div></div>';
  h+='<div class="field"><div class="field-label">MP Date</div><div class="field-value">'+escHtml(mp||'N/A')+'</div></div>';
  h+='<div class="field"><div class="field-label">BTO Ready</div><div class="field-value">'+escHtml(bto||'N/A')+'</div></div>';
  h+='<div class="field"><div class="field-label">NPM</div><div class="field-value">'+escHtml(r.npm||'N/A')+'</div></div>';
  h+='<div class="field"><div class="field-label">CPU</div><div class="field-value" style="font-size:11px">'+escHtml(r.cpu||'N/A')+'</div></div>';
  h+='<div class="field"><div class="field-label">GPU</div><div class="field-value" style="font-size:11px">'+escHtml(r.gpu||'N/A')+'</div></div>';
  h+='</div>';
  h+='<div class="section"><div class="section-title">Current Status</div><div class="section-body">'+formatSectionText(r.status||'')+'</div></div>';
  h+='<div class="section"><div class="section-title">Highlight</div><div class="section-body">'+formatSectionText(r.highlight||'')+'</div></div>';
  h+='<div class="section"><div class="section-title">Timeline</div><div class="section-body" style="padding:8px 12px">'+timelineH+'</div></div>';
  return h;
}}

// ── 初始化 ──
buildMsDropdown('model');
buildMsDropdown('mkt');
renderList();
</script>
</body>
</html>'''

    os.makedirs(os.path.dirname(SEARCH_HTML_PATH), exist_ok=True)
    with open(SEARCH_HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Search HTML 已生成: {SEARCH_HTML_PATH}")


def main():
    print(f"读取 Excel: {EXCEL_PATH}")
    records = read_excel()
    print(f"共读取 {len(records)} 条记录")

    # 跳过 GitHub 比对（build_npi_1 离线版）
    # records = detect_mp_changes(records)
    changed = []
    if changed:
        pass
    if changed:
        print(f"检测到 {len(changed)} 条 MP 变动: {', '.join(r['model']+'('+r['mkt']+')' for r in changed)}")
    else:
        print("未检测到 MP 日期或 stage 变动")

    # 保存 JSON 副本
    with open(JSON_PATH, "w", encoding="utf-8") as f:
        json.dump({"records": records, "buildTime": datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")}, f, ensure_ascii=False, indent=2)
    print(f"JSON 已保存: {JSON_PATH}")

    # 嵌入 HTML
    inject_html(records)
    print(f"HTML 已更新: {HTML_PATH}")

    # 嵌入 npi_dashboard2.html
    inject_html2(records)
    print(f"HTML2 已更新: {HTML_PATH2}")


    # 生成 Excel
    xlsx_path = generate_xlsx(records)

    # 生成 NPI Search Dashboard
    build_search_html(records)

    print("完成！双击打开 HTML 即可查看 Dashboard。")


if __name__ == "__main__":
    print("Script started...")
    try:
        main()
    except Exception as e:
        import traceback
        print("\n" + "=" * 50)
        print("FATAL ERROR:")
        traceback.print_exc()
        print("=" * 50)
        import sys
        sys.exit(1)
    input("\nPress Enter to close...")
