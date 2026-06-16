"""
从 npi_data.json 导出 Excel 报表
用法: python export_excel.py
"""

import json
import os
import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
JSON_PATH = os.path.join(SCRIPT_DIR, "npi_data.json")
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "export")

DATE_KEYS = ["Kickoff", "DVT-start", "EVT-start", "MVT-start", "BTO ready", "ATS-start", "MP"]

# 表头
HEADERS = ["No.", "Model", "MKT Name", "Stage", "CPU", "GPU", "NPM",
           "Kickoff", "DVT-start", "EVT-start", "MVT-start", "BTO ready", "ATS-start", "MP",
           "Status", "Highlight"]


def export():
    # 读取数据
    with open(JSON_PATH, "r", encoding="utf-8") as f:
        data = json.load(f)
    records = data.get("records", [])
    build_time = data.get("buildTime", "")

    if not records:
        print("npi_data.json 中没有记录，跳过导出。")
        return None

    # 用 openpyxl 生成 Excel
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("缺少 openpyxl，正在安装...")
        os.system(f'"{os.sys.executable}" -m pip install openpyxl -q')
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NPI Dashboard"

    # 样式
    header_font = Font(name="Microsoft JhengHei", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Microsoft JhengHei", size=10)
    cell_align = Alignment(vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Stage 颜色映射
    stage_fills = {
        "Kickoff": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "Design": PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
        "DVT": PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid"),
        "EVT": PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid"),
        "MVT": PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid"),
        "ATS": PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid"),
        "Study": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
        "MP": PatternFill(start_color="FF69B4", end_color="FF69B4", fill_type="solid"),
    }
    stage_font = Font(name="Microsoft JhengHei", size=10, bold=True, color="FFFFFF")

    # 写标题行
    for col_idx, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写数据
    for row_idx, r in enumerate(records, 2):
        # No.
        c = ws.cell(row=row_idx, column=1, value=row_idx - 1)
        c.font = cell_font; c.alignment = center_align; c.border = thin_border

        # Model
        c = ws.cell(row=row_idx, column=2, value=r.get("model", ""))
        c.font = Font(name="Microsoft JhengHei", size=10, bold=True); c.alignment = center_align; c.border = thin_border

        # MKT
        c = ws.cell(row=row_idx, column=3, value=r.get("mkt", ""))
        c.font = cell_font; c.alignment = cell_align; c.border = thin_border

        # Stage (带颜色)
        stage = r.get("stage", "")
        c = ws.cell(row=row_idx, column=4, value=stage)
        c.font = stage_font; c.alignment = center_align; c.border = thin_border
        if stage in stage_fills:
            c.fill = stage_fills[stage]

        # CPU
        c = ws.cell(row=row_idx, column=5, value=r.get("cpu", ""))
        c.font = cell_font; c.alignment = cell_align; c.border = thin_border

        # GPU
        c = ws.cell(row=row_idx, column=6, value=r.get("gpu", ""))
        c.font = cell_font; c.alignment = cell_align; c.border = thin_border

        # NPM
        c = ws.cell(row=row_idx, column=7, value=r.get("npm", ""))
        c.font = cell_font; c.alignment = center_align; c.border = thin_border

        # 日期列 (8-14)
        for dk_idx, dk in enumerate(DATE_KEYS):
            val = r.get("dates", {}).get(dk, "")
            c = ws.cell(row=row_idx, column=8 + dk_idx, value=val)
            c.font = cell_font; c.alignment = center_align; c.border = thin_border

        # Status
        c = ws.cell(row=row_idx, column=15, value=r.get("status", ""))
        c.font = cell_font; c.alignment = center_align; c.border = thin_border

        # Highlight
        c = ws.cell(row=row_idx, column=16, value=r.get("highlight", ""))
        c.font = cell_font; c.alignment = cell_align; c.border = thin_border

    # 列宽
    col_widths = [5, 10, 35, 8, 28, 38, 10, 12, 12, 12, 12, 12, 12, 12, 8, 50]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    # 冻结窗格（冻结标题行 + 前两列）
    ws.freeze_panes = "C2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(HEADERS))}{len(records)+1}"

    # 保存
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"NPI_Dashboard_{ts}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)
    print(f"Excel 已导出: {filepath}")
    print(f"共 {len(records)} 条记录")
    return filepath


if __name__ == "__main__":
    export()
